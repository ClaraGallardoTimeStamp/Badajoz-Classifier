"""
Genera el Excel final con formato profesional, a partir del clasificado.
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

SRC = 'clasificacion_output.xlsx'
OUT = 'Puntos_de_Interes_Clasificados_servicios.xlsx'

import os
os.makedirs('outputs', exist_ok=True)

df = pd.read_excel(SRC)

# Resumen
total = len(df)
sin_cls = int(df['sin_clasificar'].sum())
confl = int(df['conflicto_semantico'].sum())
ambig = int(df['ambiguedad'].sum())
conf_media = float(df['confianza'].mean())

# Hoja resumen
resumen = pd.DataFrame({
    'Métrica': [
        'Total de registros',
        'Clasificados correctamente',
        'Sin clasificar',
        'Conflicto semántico (ganador fuera del mapping)',
        'Ambigüedad (2+ candidatos con score similar)',
        'Confianza media',
        '% con confianza ≥ 0.7',
    ],
    'Valor': [
        total,
        total - sin_cls,
        sin_cls,
        confl,
        ambig,
        round(conf_media, 3),
        f"{(df['confianza']>=0.7).mean()*100:.1f}%",
    ]
})

# Distribución
dist = df['categoria_final'].value_counts().reset_index()
dist.columns = ['categoria_final', 'recuento']

# Matriz de transformaciones categoría_input -> categoría_final
piv = (df.groupby(['categoria', 'categoria_final'])
         .size().reset_index(name='n')
         .sort_values(['categoria','n'], ascending=[True, False]))

with pd.ExcelWriter(OUT, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Clasificacion', index=False)
    resumen.to_excel(writer, sheet_name='Resumen', index=False)
    dist.to_excel(writer, sheet_name='Distribucion_final', index=False)
    piv.to_excel(writer, sheet_name='Mapeo_origen_destino', index=False)

# Formato
wb = load_workbook(OUT)
thin = Side(border_style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', start_color='305496')
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_sheet(ws, col_widths=None, wrap_cols=None):
    wrap_cols = wrap_cols or []
    # Header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    # Body
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
            cell.border = border
            if cell.column_letter in wrap_cols:
                cell.alignment = left
            else:
                cell.alignment = Alignment(vertical='center')
    # Widths
    if col_widths:
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

# Hoja Clasificacion
ws = wb['Clasificacion']
widths = {
    'A': 42,  # Nombre
    'B': 22,  # categoria
    'C': 50,  # descripcion
    'D': 26,  # categoria_final
    'E': 24,  # Nivel 1
    'F': 28,  # Nivel 2
    'G': 26,  # Nivel 3
    'H': 24,  # Nivel 4
    'I': 11,  # confianza
    'J': 14,  # conflicto
    'K': 12,  # ambiguedad
    'L': 13,  # sin_clasificar
    'M': 55,  # explicacion
}
style_sheet(ws, widths, wrap_cols=['A','C','D','E','F','G','H','M'])

# Formato condicional en confianza
from openpyxl.formatting.rule import ColorScaleRule
ws.conditional_formatting.add(
    f'I2:I{ws.max_row}',
    ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                   mid_type='num', mid_value=0.6, mid_color='FFEB84',
                   end_type='num', end_value=1.0, end_color='63BE7B')
)

# Sin clasificar en rojo
red_fill = PatternFill('solid', start_color='FFD9D9')
yellow_fill = PatternFill('solid', start_color='FFF6C2')
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    sin_cls_cell = row[11]  # col L
    confl_cell = row[9]     # col J
    if sin_cls_cell.value is True:
        for c in row: c.fill = red_fill
    elif confl_cell.value is True:
        for c in row: c.fill = yellow_fill

# Hoja Resumen
ws = wb['Resumen']
style_sheet(ws, {'A': 55, 'B': 20})

# Hoja Distribución
ws = wb['Distribucion_final']
style_sheet(ws, {'A': 35, 'B': 12})

# Hoja Mapeo
ws = wb['Mapeo_origen_destino']
style_sheet(ws, {'A': 28, 'B': 30, 'C': 10})

wb.save(OUT)
print(f"Excel final guardado en: {OUT}")

# Reporte breve
print("\n" + "="*60)
print("RESUMEN EJECUTIVO")
print("="*60)
print(f"Total registros:          {total}")
print(f"Clasificados:             {total-sin_cls} ({(total-sin_cls)/total*100:.1f}%)")
print(f"Sin clasificar:           {sin_cls}")
print(f"Conflicto semántico:      {confl} ({confl/total*100:.1f}%)")
print(f"Ambigüedad:               {ambig}")
print(f"Confianza media:          {conf_media:.3f}")
print(f"Con confianza ≥ 0.7:      {(df['confianza']>=0.7).sum()} ({(df['confianza']>=0.7).mean()*100:.1f}%)")
